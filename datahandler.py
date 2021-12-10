from openpyxl import load_workbook
from panels import *
import pandas as pd

# Load existing Excel file, has to follow template
def load_file(file_loc):
    workbook = load_workbook(filename=file_loc)
    snames = workbook.sheetnames
    snames = snames[:len(snames)-2]

    panel_list = []
    for name in snames:
        ws = workbook[name]
        p = Panel(name, 42)
        p.set_voltages(ws['B2'].value, ws['B4'].value)

        # Process left side of panel
        for val in ws['I10:I30']:
            name = str(val[0].value)
            if 'PANEL' in name:
                row_col = [val[0].row, val[0].column]

                link = str.split(name, ' ')[1]
                p.append_connection(link)
                s_a = "='" + link + "'!M31*1000"
                s_b = "='" + link + "'!O31*1000"
                s_c = "='" + link + "'!Q31*1000"
                ws.cell(row=row_col[0], column=row_col[1]-2).value = s_a
                ws.cell(row=row_col[0]+1, column=row_col[1]-2).value = s_b
                ws.cell(row=row_col[0]+2, column=row_col[1]-2).value = s_c

        # Process right side of panel
        for val in ws['V10:V30']:
            name = str(val[0].value)
            if 'PANEL' in name:
                row_col = [val[0].row, val[0].column]

                link = str.split(name, ' ')[1]
                p.append_connection(link)
                s_a = "='" + link + "'!M31*1000"
                s_b = "='" + link + "'!O31*1000"
                s_c = "='" + link + "'!Q31*1000"
                ws.cell(row=row_col[0], column=row_col[1]+8).value = s_a
                ws.cell(row=row_col[0]+1, column=row_col[1]+8).value = s_b
                ws.cell(row=row_col[0]+2, column=row_col[1]+8).value = s_c
        workbook.save('excel_sheets/results.xlsx')

        panel_list.append(p)
    panel_dict = {}
    for p in panel_list:
        vals = pd.read_excel(file_loc, sheet_name=p.name)
        con_load = float(vals.iloc[32]['Unnamed: 12']) * 1000
        p.load_va = con_load
        panel_dict[p.name] = p

    # Generate list of transformers
    xfmr_list = []
    for p2 in panel_list:
        if len(p2.connections) > 0:
            count = 1
            for con in p2.connections:
                n = 'XMFR_' + str(count)
                xfmr = Transformer(str(n), str(p2.name), str(con),
                                   p2.ll_volt, panel_dict[con].ll_volt, panel_dict[con].load_va)
                xfmr_list.append(xfmr)
                count = count + 1

    return panel_dict, xfmr_list


# Update the transformer schedule in the Excel File based on connections between panels
def write_xfmr_sched(xfmr_list):
    workbook = load_workbook(filename='excel_sheets/results.xlsx')
    ws = workbook['xfmr_sched']
    row = 4
    with open('rating_information/transformers.txt') as f:
        lines = f.readlines()
    xfmr_standard = []
    for line in lines:
        line = line.replace('\n', '')
        line = float(line)*1000
        xfmr_standard.append(line)

    for xfmr in xfmr_list:
        #determine next standard xfmr based on 125% future load overhead
        xfmr_val = float(xfmr.load_va)*1.25

        for s in xfmr_standard:
            if xfmr_val < s:
                ws.cell(row=row, column=7).value = s/1000
                break

        ws.cell(row=row, column=1).value = row-3
        ws.cell(row=row, column=2).value = xfmr.name
        ws.cell(row=row, column=3).value = xfmr.input_con
        ws.cell(row=row, column=4).value = xfmr.output_con
        ws.cell(row=row, column=5).value = 'delta-wye'
        ws.cell(row=row, column=6).value = xfmr.load_va/1000
        ws.cell(row=row, column=8).value = xfmr.input_v
        ws.cell(row=row, column=9).value = xfmr.output_v
        row = row + 1

    workbook.save('excel_sheets/results.xlsx')


# Update feeder schedule based on inputs and outputs to transformers
def write_feeder_sched(xfmr_list, net, wire_type='CU'):
    workbook = load_workbook(filename='excel_sheets/results.xlsx')
    ws = workbook['feeder_sched']
    phase_list, ground_list = load_wire_info()
    net_list = net.nodes

    phase_check = []
    ground_idx = 0
    if wire_type == 'CU':
        ground_idx = 1
        for ph in phase_list:
            phase_check.append(ph[2])
    else:
        ground_idx = 2
        for ph in phase_list:
            phase_check.append(ph[2])

    ground_check = []
    for gr in ground_list:
        ground_check.append(gr[0])
    print(ground_check)

    inc = 4
    for xfmr in xfmr_list:
        inp_amps = xfmr.load_va/xfmr.input_v
        oup_amps = xfmr.load_va/xfmr.output_v

        in_phase_gauge = check_val(inp_amps, phase_list, phase_check, 1)
        in_ground_gauge = check_val(inp_amps, ground_list, ground_check, 1)
        out_phase_gauge = check_val(oup_amps, phase_list, phase_check, 1)
        out_ground_gauge = check_val(oup_amps, ground_list, ground_check, 1)

        in_desc = '(3) ' + in_phase_gauge + ', (1) ' + in_ground_gauge + ' G'
        out_desc = '(4) ' + out_phase_gauge + ', (1) ' + out_ground_gauge + ' G'

        ws.cell(row=inc, column=1).value = inc - 3
        ws.cell(row=inc, column=2).value = xfmr.input_con + '->' + xfmr.name
        ws.cell(row=inc, column=3).value = 1
        ws.cell(row=inc, column=4).value = in_desc
        ws.cell(row=inc, column=5).value = wire_type
        ws.cell(row=inc, column=6).value = 'EMT'

        ws.cell(row=inc+1, column=1).value = inc - 2
        ws.cell(row=inc+1, column=2).value = xfmr.name + '->' + xfmr.output_con
        ws.cell(row=inc+1, column=3).value = 1
        ws.cell(row=inc+1, column=4).value = out_desc
        ws.cell(row=inc+1, column=5).value = wire_type
        ws.cell(row=inc+1, column=6).value = 'EMT'

        inc = inc + 2

    workbook.save('excel_sheets/results.xlsx')


def check_val(val, o_list, c_list, idx):
    count = 0
    out = ''
    for c in c_list:
        if val < float(c):
            out = o_list[count][idx]
            break
        count = count + 1
    if len(o_list[0]) == 4:
        return o_list[count][0] + out
    return out


def load_wire_info():
    with open('rating_information/phase_wire_sizing.txt') as f:
        phase = f.readlines()
    with open('rating_information/ground_wire_sizing.txt') as f:
        ground = f.readlines()
    phase_list = []
    ground_list = []
    for p in phase:
        p = p.replace('\n', '')
        phase_list.append(str.split(p, ' '))
    for g in ground:
        g = g.replace('\n', '')
        ground_list.append(str.split(g, ' '))
    return phase_list, ground_list

